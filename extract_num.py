from matplotlib import pyplot as plt
import numpy as np
import cv2
from openpyxl import load_workbook


class ExtractNum:
    def __init__(self, image_path, temp_path, target_path, i):
        # Load the image from the file
        self.image_path = image_path
        self.temp_path = temp_path
        self.i = i
        self.target_path = target_path

        self.extract_red(self.image_path, self.temp_path)
        self.image = cv2.imread(self.temp_path)
        self.digits, self.location = self.extracted_digits(self.image)
        self.mnist_digits = self.convert2mnist(self.digits)
        #       self.display()
        self.normalized_num = self.convert21d(self.mnist_digits)

    def extracted_digits(self, image):
        gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        # Apply thresholding to get the digits to stand out
        _, binary_image = cv2.threshold(gray_image, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        # Find contours - these should correspond to the digits
        contours, _ = cv2.findContours(binary_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        # Filter contours by size to eliminate any small noise
        digit_contours = [c for c in contours if cv2.contourArea(c) > 100]
        # Sort the contours from left to right
        digit_contours = sorted(digit_contours, key=lambda ctr: cv2.boundingRect(ctr)[0])
        extracted_digits = []
        location = []
        for contour in digit_contours:
            # Get the bounding box of the contour
            x, y, w, h = cv2.boundingRect(contour)
            # Crop the image to the bounding box size
            digit_img = binary_image[y:y + h, x:x + w]
            # Append to our list of extracted digits
            extracted_digits.append(digit_img)
            location.append([x, y, w, h])

        return extracted_digits, location


    def convert2mnist(self, images):
        mnist_format_images = []
        for img in images:
            # Resize image to 20x20 to fit in the 28x28 MNIST format
            resized_img = cv2.resize(img, (20, 20), interpolation=cv2.INTER_AREA)
            # Add padding to make image 28x28
            padded_img = cv2.copyMakeBorder(resized_img, 4, 4, 4, 4, cv2.BORDER_CONSTANT, value=[0, 0, 0])
            # Normalize the image pixel values to 0-1
            normalized_img = padded_img
            mnist_format_images.append(normalized_img)

        return mnist_format_images

    def display(self, images):
        for image in range(len(images)):
            plt.imshow(image, cmap='gray')
            plt.title('MNIST Format Digit')
            plt.axis('off')
            plt.show()

    def convert21d(self, mnist_digits):
        # Convert the list of 28x28 images to a numpy array
        mnist_digits_array = np.array(mnist_digits)
        # Rescale the values back to 0-255 and convert to integers
        mnist_digits_array = (mnist_digits_array).astype('uint8')
        # Flatten each 28x28 image into a 784-length vector
        return np.array([digit.reshape(784, ) for digit in mnist_digits_array])
        # Now X is an (m x 784) dimension ndarray of integers between 0-255

    def convert2nums(self, prediction, precision=0.2):
        n = len(prediction)
        res = []
        i = 0
        while i < n:
            if i != n - 1:
                x1, y1, w1, h1 = self.location[i]
                x2, y2, w2, h2 = self.location[i + 1]
                w1, h2 = max(w1, h1), max(w1, h1)
                if x1 + w1 * (1 + precision) >= x2 - (max(w2, h2) - w2) / 2 * (1 + precision):
                    res.append(str(prediction[i]) + str(prediction[i + 1]))
                    i += 2
                    continue
            res.append(str(prediction[i]))
            i += 1

        length = len(res)
        if length > 6:
            res = res[:6]
        elif length < 6:
            res += [0] * (6 - length)

        formation = [range(16), range(16), range(11), range(21), range(21), range(21)]

        for i in range(6):
            if int(res[i]) not in formation[i]:
                res[i] = str(max(formation[i]))
        return res

    def extract_red(self, input_path, out_path):
        image = cv2.imread(input_path)
        image_hsv = cv2.cvtColor(image, cv2.COLOR_BGR2HSV)
        # Define the lower and upper bounds for the red color in HSV
        # Red has a hue of 0 or 180 degrees, depending on how it's represented.
        # We'll create two masks, one for the lower red and one for the higher red hue values,
        # and then combine them to get the full range of red.

        # Lower mask (0-10)
        lower_red = np.array([0, 50, 50])
        upper_red = np.array([10, 255, 255])
        mask0 = cv2.inRange(image_hsv, lower_red, upper_red)

        # Upper mask (170-180)
        lower_red = np.array([100, 50, 50])
        upper_red = np.array([200, 255, 255])
        mask1 = cv2.inRange(image_hsv, lower_red, upper_red)

        # Join the masks
        mask = mask0 + mask1
        # Bitwise-AND mask and original image to isolate the red components
        isolated_red = cv2.bitwise_and(image, image, mask=mask)
        # Convert to grayscale to simplify the image further, thresholding it to make the digits more distinct.
        gray_red = cv2.cvtColor(isolated_red, cv2.COLOR_BGR2GRAY)
        _, thresh_red = cv2.threshold(gray_red, 1, 255, cv2.THRESH_BINARY)
        # Show the threshold image which will be used for digit recognition
        # We'll save the image so it can be used externally.
        cv2.imwrite(out_path, 255 - thresh_red)

    def excel_writer(self, data):
        wb = load_workbook(self.target_path)
        ws = wb['分项分数和散点图']
        i = self.i + 1

        formulation = {'总分': f'=ROUND(H{i}*0.7+AC{i}*0.3,0)',
                       '总分目标达成度': f'=ROUND(F{i}/100,2)',
                       '考试分': f'=I{i}+J{i}+K{i}+N{i}+Q{i}+R{i}',
                       '第一大题': data[0], '第二大题': data[1], '第三大题': data[2],
                       '第四大题': data[3], '第五大题': data[4], '第六大题': data[5],
                       '课程目标1得分': f'=ROUND((I{i}+J{i}+K{i})*0.7+(V{i}*0.3)*0.5+(Z{i}*0.3)*0.5,0)',
                       '课程目标1达成度': f'=ROUND(L{i}/37,2)',
                       '课程目标2得分': f'=ROUND(N{i}*0.7+W{i}*0.3*0.5+AA{i}*0.3*0.5,0)',
                       '课程目标2达成度': f'=ROUND(O{i}/26,2)',
                       '课程目标3得分': f'=ROUND((Q{i}+R{i})*0.7+X{i}*0.3*0.5+AB{i}*0.3*0.5,0)',
                       '课程目标3达成度': f'=ROUND(S{i}/37,2)'
                       }

        # Define headers you need to work with

        header_columns = {}
        for col_index in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col_index).value
            header_columns[header] = col_index

        for header in header_columns:
            if header in formulation:
                ws.cell(row=i, column=header_columns[header]).value = formulation[header]
        wb.save(self.target_path)
